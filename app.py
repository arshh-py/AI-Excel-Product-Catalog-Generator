import hashlib
import html
import json
import logging
import os
import re
import sqlite3
import struct
import sys
import threading
import time
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, unquote, urljoin, urlparse
import webbrowser

import pandas as pd
import requests
from ddgs import DDGS
from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from models import IMAGE_CACHE_SCHEMA, IMPORT_JOBS_SCHEMA, PRODUCTS_SCHEMA, UPLOAD_BATCHES_SCHEMA


if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", BASE_DIR))
else:
    BASE_DIR = Path(__file__).resolve().parent
    RESOURCE_DIR = BASE_DIR

TEMPLATE_DIR = RESOURCE_DIR / "templates"
STATIC_DIR = RESOURCE_DIR / "static"
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
IMAGE_CACHE_DIR = DATA_DIR / "image_cache"
DATABASE_PATH = DATA_DIR / "inventory.sqlite3"
IMAGE_SEARCH_LOG_PATH = DATA_DIR / "image_search.log"
DATABASE_TIMEOUT_SECONDS = 60
HTTP_TIMEOUT = (3, 8)
FAST_HTTP_TIMEOUT = (2, 4)
MAX_IMAGE_RESULTS_PER_QUERY = 12
MAX_WEB_RESULTS_PER_QUERY = 6
MAX_PAGES_PER_WEB_QUERY = 2
MAX_CANDIDATES_PER_SOURCE = 12
MAX_DOWNLOADS_PER_SOURCE = 8
FAST_IMAGE_RESULTS_PER_QUERY = 6
FAST_WEB_RESULTS_PER_QUERY = 4
FAST_PAGES_PER_WEB_QUERY = 1
FAST_CANDIDATES_PER_SOURCE = 4
FAST_DOWNLOADS_PER_SOURCE = 2

ALLOWED_EXTENSIONS = {".xlsx"}
PLACEHOLDER_IMAGE = "placeholder.svg"
STOP_WORDS = {
    "and",
    "or",
    "for",
    "with",
    "without",
    "the",
    "a",
    "an",
    "in",
    "of",
    "to",
    "by",
    "from",
    "new",
    "latest",
    "best",
    "combo",
    "pack",
    "set",
    "pcs",
    "pc",
    "piece",
    "pieces",
}

OPTION_WORDS_PATTERN = re.compile(
    r"\b("
    r"\d+\s?\+\s?\d+\s?gb|"
    r"\d+\s?gb\s?ram|"
    r"\d+\s?(?:gb|tb|mb|mah|w|hz|inch|inches|cm|mm)|"
    r"black|white|blue|green|red|yellow|silver|gold|grey|gray|purple|pink|orange|"
    r"midnight|starlight|graphite|charcoal|cream|beige|brown|transparent|"
    r"ram|rom|storage|memory|color|colour|variant|edition|renewed|refurbished|"
    r"with|without|combo|pack|set|new|latest"
    r")\b",
    re.IGNORECASE,
)

MODEL_PATTERN = re.compile(r"\b[A-Z]{1,6}[- ]?\d[A-Z0-9-]{1,}\b|\b[A-Z0-9]{2,}[-/][A-Z0-9-]{2,}\b")

KNOWN_BRANDS = {
    "acer",
    "amazon",
    "apple",
    "asus",
    "boat",
    "bosch",
    "canon",
    "dell",
    "dyson",
    "fire-boltt",
    "godrej",
    "google",
    "haier",
    "havells",
    "hp",
    "ifb",
    "jbl",
    "lenovo",
    "lg",
    "mi",
    "microsoft",
    "motorola",
    "noise",
    "oneplus",
    "oppo",
    "panasonic",
    "philips",
    "realme",
    "redmi",
    "samsung",
    "sony",
    "vivo",
    "voltas",
    "whirlpool",
    "xiaomi",
}

SOURCE_PLANS = [
    {
        "name": "Amazon product pages",
        "mode": "web",
        "queries": ["{query} site:amazon.in", "{query} site:amazon.com"],
        "domains": ("amazon.in", "amazon.com"),
        "max_pages": 2,
    },
    {
        "name": "Amazon image CDN",
        "mode": "images",
        "queries": [
            "{query} site:m.media-amazon.com",
            "{query} site:images-na.ssl-images-amazon.com",
            "{query} amazon image",
        ],
        "required_hosts": ("media-amazon.com", "images-amazon.com", "ssl-images-amazon.com"),
    },
    {"name": "Google Images", "mode": "images", "queries": ["{query} product image", "{query} front image"]},
    {"name": "Google Web Search", "mode": "web", "queries": ["{query} product image", "{query} buy online"], "max_pages": 2},
    {
        "name": "Official manufacturer websites",
        "mode": "web",
        "queries": ["{brand_model} official product image", "{brand_model} official website"],
        "exclude_domains": (
            "amazon.",
            "flipkart.",
            "reliancedigital.",
            "croma.",
            "vijaysales.",
            "ebay.",
            "walmart.",
        ),
        "max_pages": 2,
    },
    {"name": "Flipkart", "mode": "web", "queries": ["{query} site:flipkart.com"], "domains": ("flipkart.com",), "max_pages": 2},
    {
        "name": "Reliance Digital",
        "mode": "web",
        "queries": ["{query} site:reliancedigital.in"],
        "domains": ("reliancedigital.in",),
        "max_pages": 2,
    },
    {"name": "Croma", "mode": "web", "queries": ["{query} site:croma.com"], "domains": ("croma.com",), "max_pages": 2},
    {
        "name": "Vijay Sales",
        "mode": "web",
        "queries": ["{query} site:vijaysales.com"],
        "domains": ("vijaysales.com",),
        "max_pages": 2,
    },
    {"name": "Bing Images", "mode": "images", "queries": ["{query} product image", "{query} buy image"]},
    {"name": "DuckDuckGo Images", "mode": "images", "queries": ["{query} product image", "{query} front view"]},
    {"name": "eBay", "mode": "web", "queries": ["{query} site:ebay.com", "{query} site:ebay.in"], "domains": ("ebay.com", "ebay.in"), "max_pages": 2},
    {"name": "Walmart", "mode": "web", "queries": ["{query} site:walmart.com"], "domains": ("walmart.com",), "max_pages": 2},
    {
        "name": "Trusted shopping websites",
        "mode": "web",
        "queries": [
            "{query} buy online product image",
            "{query} shop product image",
            "{query} price product image",
        ],
        "max_pages": 2,
    },
]

BULK_IMPORT_SOURCE_PLANS = [
    SOURCE_PLANS[1],   # Amazon image CDN
    SOURCE_PLANS[2],   # Google Images-style image search
    SOURCE_PLANS[10],  # DuckDuckGo Images-style image search
    SOURCE_PLANS[9],   # Bing Images-style image search
]


app = Flask(__name__, template_folder=str(TEMPLATE_DIR), static_folder=str(STATIC_DIR))
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "local-inventory-dev-key")
app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024

image_search_logger = logging.getLogger("image_search")
image_search_logger.setLevel(logging.INFO)
image_search_logger.propagate = False


def ensure_directories():
    """Create local storage folders the app needs at runtime."""
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    IMAGE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    if not image_search_logger.handlers:
        handler = logging.FileHandler(IMAGE_SEARCH_LOG_PATH, encoding="utf-8")
        handler.setFormatter(logging.Formatter("%(asctime)s %(message)s"))
        image_search_logger.addHandler(handler)


def log_image_search(event, **fields):
    ensure_directories()
    payload = {"event": event, **fields}
    image_search_logger.info(json.dumps(payload, ensure_ascii=False, sort_keys=True))


def get_db():
    conn = sqlite3.connect(
        DATABASE_PATH,
        timeout=DATABASE_TIMEOUT_SECONDS,
        isolation_level=None,
    )
    conn.row_factory = sqlite3.Row
    conn.execute(f"PRAGMA busy_timeout = {DATABASE_TIMEOUT_SECONDS * 1000}")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    ensure_directories()
    with get_db() as conn:
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute("PRAGMA synchronous = NORMAL")
        conn.execute(PRODUCTS_SCHEMA)
        conn.execute(IMAGE_CACHE_SCHEMA)
        conn.execute(UPLOAD_BATCHES_SCHEMA)
        conn.execute(IMPORT_JOBS_SCHEMA)
        conn.execute(
            """
            UPDATE import_jobs
            SET status = ?, error_message = ?, current_product = ?, completed_at = ?, updated_at = ?
            WHERE status IN (?, ?)
            """,
            (
                "error",
                "App was restarted before this import finished. Please upload again.",
                "Import stopped",
                now_iso(),
                now_iso(),
                "queued",
                "running",
            ),
        )


def allowed_excel_file(filename):
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def normalize_product_name(value):
    if value is None or pd.isna(value):
        return ""
    name = re.sub(r"\s+", " ", str(value)).strip()
    return name


def product_hash(product_name):
    normalized = product_name.casefold().strip()
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:24]


def now_iso():
    return datetime.utcnow().isoformat(timespec="seconds")


def get_price_text(value):
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_column_name(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().casefold())


def find_column(columns, expected):
    normalized = {normalize_column_name(col): col for col in columns}
    return normalized.get(normalize_column_name(expected))


def find_first_column(columns, candidates):
    normalized = {normalize_column_name(col): col for col in columns}
    for candidate in candidates:
        column = normalized.get(normalize_column_name(candidate))
        if column is not None:
            return column
    return None


def numeric_value(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if cleaned in {"", "-", ".", "-."}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def format_price_value(value):
    if value is None:
        return ""
    rounded = round(float(value), 2)
    if rounded.is_integer():
        return str(int(rounded))
    return f"{rounded:.2f}".rstrip("0").rstrip(".")


def detect_price_columns(columns):
    unit_price_col = find_first_column(
        columns,
        (
            "PRICE/PC",
            "PRICE PER PC",
            "PRICE/PCS",
            "PRICE PER PCS",
            "PRICE PER PIECE",
            "PER PIECE PRICE",
            "UNIT PRICE",
            "UNITPRICE",
            "RATE",
        ),
    )
    quantity_col = find_first_column(columns, ("QTY", "QUANTITY", "PCS", "PIECES"))
    total_price_col = find_first_column(
        columns,
        (
            "TOTAL/PRICE",
            "TOTAL PRICE",
            "TOTALPRICE",
            "TOTAL AMOUNT",
            "TOTALAMOUNT",
            "AMOUNT",
            "T ASP",
            "TOTAL ASP",
            "TOTALASP",
        ),
    )
    fallback_price_col = find_first_column(columns, ("ASP", "PRICE", "SALE PRICE", "SELLING PRICE", "MRP"))
    return {
        "unit_price": unit_price_col,
        "quantity": quantity_col,
        "total_price": total_price_col,
        "fallback_price": fallback_price_col,
    }


def get_unit_price_text(row, price_columns):
    unit_price_col = price_columns.get("unit_price")
    if unit_price_col is not None:
        unit_price = numeric_value(row.get(unit_price_col))
        if unit_price is not None:
            return format_price_value(unit_price)
        return get_price_text(row.get(unit_price_col))

    total_price_col = price_columns.get("total_price")
    quantity_col = price_columns.get("quantity")
    total_price = numeric_value(row.get(total_price_col)) if total_price_col is not None else None
    quantity = numeric_value(row.get(quantity_col)) if quantity_col is not None else None
    if total_price is not None and quantity and quantity > 0:
        return format_price_value(total_price / quantity)

    fallback_price_col = price_columns.get("fallback_price")
    if fallback_price_col is not None:
        fallback_price = numeric_value(row.get(fallback_price_col))
        if fallback_price is not None:
            return format_price_value(fallback_price)
        return get_price_text(row.get(fallback_price_col))

    if total_price is not None:
        return format_price_value(total_price)
    return ""


def clean_query_text(value):
    return re.sub(r"\s+", " ", value).strip(" -|,;:/")


def simplify_product_name(product_name):
    simplified = OPTION_WORDS_PATTERN.sub(" ", product_name)
    simplified = re.sub(r"\([^)]*\)|\[[^]]*]", " ", simplified)
    return clean_query_text(simplified)


def detect_model_numbers(product_name):
    models = []
    for match in MODEL_PATTERN.finditer(product_name):
        model = match.group(0).strip()
        if model.casefold().startswith(("ram ", "rom ", "storage ")):
            continue
        models.append(model)
    return list(dict.fromkeys(models))


def detect_brand(product_name):
    tokens = re.findall(r"[a-z0-9-]+", product_name.casefold())
    for token in tokens:
        if token in KNOWN_BRANDS:
            return token
    return tokens[0] if tokens else ""


def build_search_queries(product_name):
    """Use full title first, then increasingly broad fallbacks."""
    cleaned = clean_query_text(product_name)
    short = re.split(r"\||,| - | with | for ", cleaned, maxsplit=1, flags=re.IGNORECASE)[0]
    short = clean_query_text(short)
    simplified = simplify_product_name(cleaned)
    brand = detect_brand(cleaned)
    models = detect_model_numbers(cleaned)

    queries = [cleaned]
    if short and short.casefold() != cleaned.casefold():
        queries.append(short)
    if simplified and simplified.casefold() not in {item.casefold() for item in queries}:
        queries.append(simplified)
    for model in models:
        queries.append(model)
        if brand:
            queries.append(f"{brand} {model}")
    if brand and simplified and not simplified.casefold().startswith(f"{brand} "):
        queries.append(f"{brand} {simplified}")

    official_name = detect_official_product_name(cleaned)
    if official_name:
        queries.append(official_name)

    return list(dict.fromkeys(clean_query_text(query) for query in queries if clean_query_text(query)))


def detect_official_product_name(product_name):
    cleaned = clean_query_text(product_name)
    separators = ("|", " - ", ", with ", " with ", " for ")
    for separator in separators:
        if separator in cleaned:
            candidate = clean_query_text(cleaned.split(separator, 1)[0])
            if len(candidate) >= 8:
                return candidate
    return None


def source_query_text(template, query, product_name):
    brand = detect_brand(product_name)
    models = detect_model_numbers(product_name)
    brand_model = f"{brand} {models[0]}" if brand and models else query
    return template.format(
        query=query,
        quoted_query=f'"{query}"',
        brand=brand,
        model=models[0] if models else query,
        brand_model=brand_model,
    )


def title_tokens(text):
    tokens = re.findall(r"[a-z0-9]+", text.casefold())
    return {token for token in tokens if len(token) > 1 and token not in STOP_WORDS}


def score_image_result(product_name, result, source_name=""):
    product_tokens = title_tokens(product_name)
    result_text = " ".join(
        str(result.get(key) or "")
        for key in ("title", "source", "url", "image", "thumbnail", "page_url", "alt")
    )
    result_tokens = title_tokens(result_text)
    if not product_tokens:
        return 0

    overlap = len(product_tokens & result_tokens)
    score = overlap / len(product_tokens)
    source_text = result_text.casefold()

    if product_name.casefold() in source_text:
        score += 0.8
    if "amazon" in source_text or "flipkart" in source_text:
        score += 0.35
    if "official" in source_name.casefold():
        score += 0.15
    if "product" in source_text:
        score += 0.1
    if result.get("image"):
        score += 0.15

    width = result.get("width") or 0
    height = result.get("height") or 0
    try:
        if int(width) >= 250 and int(height) >= 250:
            score += 0.1
    except (TypeError, ValueError):
        pass

    return score


def extension_from_response(response, url):
    content_type = response.headers.get("Content-Type", "").lower()
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    if "gif" in content_type:
        return ".gif"
    if "jpeg" in content_type or "jpg" in content_type:
        return ".jpg"

    suffix = Path(urlparse(url).path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        return suffix
    return ".jpg"


def request_headers():
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36"
        ),
        "Accept-Language": "en-IN,en;q=0.9",
    }


def image_search_limits(deep_search=False):
    if deep_search:
        return {
            "timeout": HTTP_TIMEOUT,
            "image_results": MAX_IMAGE_RESULTS_PER_QUERY,
            "web_results": MAX_WEB_RESULTS_PER_QUERY,
            "pages": MAX_PAGES_PER_WEB_QUERY,
            "candidates": MAX_CANDIDATES_PER_SOURCE,
            "downloads": MAX_DOWNLOADS_PER_SOURCE,
            "source_plans": SOURCE_PLANS,
        }
    return {
        "timeout": FAST_HTTP_TIMEOUT,
        "image_results": FAST_IMAGE_RESULTS_PER_QUERY,
        "web_results": FAST_WEB_RESULTS_PER_QUERY,
        "pages": FAST_PAGES_PER_WEB_QUERY,
        "candidates": FAST_CANDIDATES_PER_SOURCE,
        "downloads": FAST_DOWNLOADS_PER_SOURCE,
        "source_plans": BULK_IMPORT_SOURCE_PLANS,
    }


def image_metadata(path):
    data = path.read_bytes()
    if data.startswith(b"\x89PNG\r\n\x1a\n") and len(data) >= 24:
        width, height = struct.unpack(">II", data[16:24])
        return {"width": width, "height": height, "format": "PNG"}
    if data.startswith((b"GIF87a", b"GIF89a")) and len(data) >= 10:
        width, height = struct.unpack("<HH", data[6:10])
        return {"width": width, "height": height, "format": "GIF"}
    if data.startswith(b"\xff\xd8"):
        offset = 2
        while offset + 9 < len(data):
            if data[offset] != 0xFF:
                offset += 1
                continue
            marker = data[offset + 1]
            offset += 2
            if marker in (0xD8, 0xD9):
                continue
            if offset + 2 > len(data):
                break
            segment_length = struct.unpack(">H", data[offset : offset + 2])[0]
            if marker in {
                0xC0,
                0xC1,
                0xC2,
                0xC3,
                0xC5,
                0xC6,
                0xC7,
                0xC9,
                0xCA,
                0xCB,
                0xCD,
                0xCE,
                0xCF,
            }:
                height, width = struct.unpack(">HH", data[offset + 3 : offset + 7])
                return {"width": width, "height": height, "format": "JPEG"}
            offset += segment_length
    if data.startswith(b"RIFF") and data[8:12] == b"WEBP" and len(data) >= 30:
        chunk = data[12:16]
        if chunk == b"VP8X" and len(data) >= 30:
            width = int.from_bytes(data[24:27], "little") + 1
            height = int.from_bytes(data[27:30], "little") + 1
            return {"width": width, "height": height, "format": "WEBP"}
        if chunk == b"VP8 " and len(data) >= 30:
            width, height = struct.unpack("<HH", data[26:30])
            return {"width": width & 0x3FFF, "height": height & 0x3FFF, "format": "WEBP"}
        if chunk == b"VP8L" and len(data) >= 25:
            bits = int.from_bytes(data[21:25], "little")
            width = (bits & 0x3FFF) + 1
            height = ((bits >> 14) & 0x3FFF) + 1
            return {"width": width, "height": height, "format": "WEBP"}
    raise ValueError("Downloaded file is not a readable image.")


def score_downloaded_image(product_name, candidate, metadata, byte_size):
    width = metadata["width"]
    height = metadata["height"]
    area = width * height
    score = score_image_result(product_name, candidate, candidate.get("source", ""))

    if width >= 500 and height >= 500:
        score += 1.2
    elif width >= 300 and height >= 300:
        score += 0.7
    elif width < 180 or height < 180:
        score -= 2.0

    score += min(area / 1_000_000, 2.0)

    image_url = candidate.get("url", "").casefold()
    page_url = candidate.get("page_url", "").casefold()
    text = " ".join([image_url, page_url, candidate.get("title", "").casefold()])

    if any(token in text for token in ("sprite", "logo", "icon", "banner", "advert", "ads.", "placeholder")):
        score -= 2.5
    if any(token in text for token in ("watermark", "youtube", "facebook", "instagram")):
        score -= 1.5
    if any(token in text for token in ("m.media-amazon.com", "images-na.ssl-images-amazon.com")):
        score += 0.5
    if width == height:
        score += 0.25
    if byte_size > 40_000:
        score += 0.25

    return score


def candidate_url_quality(url):
    text = url.casefold()
    score = 0
    if "sl1500" in text or "ul1500" in text:
        score += 3
    elif "sl1000" in text or "ul1000" in text:
        score += 2
    elif re.search(r"_[sx][yxl]?\d{3,4}_", text):
        score += 1
    if any(token in text for token in ("sprite", "logo", "icon", "banner", "advert", "placeholder")):
        score -= 5
    return score


def canonical_image_key(url):
    parsed = urlparse(url)
    key = f"{parsed.netloc.casefold()}{parsed.path}"
    if "media-amazon.com" in parsed.netloc.casefold():
        key = re.sub(r"\._[A-Z0-9,]+_\.", ".", key, flags=re.IGNORECASE)
    return key


def download_image(image_url, title_hash):
    filename, _metadata, _size = download_image_to_cache(image_url, title_hash)
    return filename


def download_image_to_cache(image_url, title_hash, suffix="", timeout=HTTP_TIMEOUT):
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36"
        )
    }
    response = requests.get(image_url, headers=headers, timeout=timeout, stream=True)
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "").lower()
    if not content_type.startswith("image/"):
        raise ValueError("The search result did not return an image.")

    ext = extension_from_response(response, image_url)
    filename = f"{title_hash}{suffix}{ext}"
    path = IMAGE_CACHE_DIR / filename

    size = 0
    with path.open("wb") as output:
        for chunk in response.iter_content(chunk_size=8192):
            if not chunk:
                continue
            size += len(chunk)
            if size > 8 * 1024 * 1024:
                raise ValueError("Image is larger than 8 MB.")
            output.write(chunk)

    if size < 1024:
        path.unlink(missing_ok=True)
        raise ValueError("Downloaded image is too small.")
    metadata = image_metadata(path)
    if metadata["width"] < 120 or metadata["height"] < 120:
        path.unlink(missing_ok=True)
        raise ValueError("Downloaded image dimensions are too small.")
    return filename, metadata, size


def normalize_result_url(url):
    if not url:
        return ""
    url = html.unescape(str(url)).strip()
    parsed = urlparse(url)
    if "uddg" in parse_qs(parsed.query):
        return unquote(parse_qs(parsed.query)["uddg"][0])
    return url


def host_matches(url, domains):
    host = urlparse(url).netloc.casefold()
    return any(domain.casefold() in host for domain in domains)


def html_attribute_values(markup, attribute):
    pattern = re.compile(rf'{attribute}\s*=\s*["\']([^"\']+)["\']', re.IGNORECASE)
    return [html.unescape(match.group(1)) for match in pattern.finditer(markup)]


def extract_json_ld_images(markup):
    images = []
    for match in re.finditer(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        markup,
        re.IGNORECASE | re.DOTALL,
    ):
        raw = html.unescape(match.group(1)).strip()
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            continue
        stack = data if isinstance(data, list) else [data]
        while stack:
            item = stack.pop()
            if isinstance(item, dict):
                image = item.get("image")
                if isinstance(image, str):
                    images.append(image)
                elif isinstance(image, list):
                    images.extend(str(value) for value in image if value)
                stack.extend(value for value in item.values() if isinstance(value, (dict, list)))
            elif isinstance(item, list):
                stack.extend(item)
    return images


def extract_images_from_page(page_url, product_name, source_name, limits):
    try:
        response = requests.get(page_url, headers=request_headers(), timeout=limits["timeout"])
        response.raise_for_status()
    except Exception as exc:
        log_image_search(
            "page_fetch_failed",
            product=product_name,
            source=source_name,
            page_url=page_url,
            error=str(exc),
        )
        return []

    content_type = response.headers.get("Content-Type", "").casefold()
    if "text/html" not in content_type:
        return []

    markup = response.text
    page_title_match = re.search(r"<title[^>]*>(.*?)</title>", markup, re.IGNORECASE | re.DOTALL)
    page_title = clean_query_text(re.sub(r"<[^>]+>", " ", html.unescape(page_title_match.group(1)))) if page_title_match else ""

    image_urls = []
    for prop in (
        "og:image",
        "og:image:secure_url",
        "twitter:image",
        "twitter:image:src",
        "image",
    ):
        image_urls.extend(
            match.group(1)
            for match in re.finditer(
                rf'<meta[^>]+(?:property|name)=["\']{re.escape(prop)}["\'][^>]+content=["\']([^"\']+)["\']',
                markup,
                re.IGNORECASE,
            )
        )
    image_urls.extend(extract_json_ld_images(markup))
    image_urls.extend(html_attribute_values(markup, "data-old-hires"))
    image_urls.extend(html_attribute_values(markup, "data-a-dynamic-image"))
    image_urls.extend(html_attribute_values(markup, "src"))

    candidates = []
    seen = set()
    for raw_url in image_urls:
        if raw_url.strip().startswith("{"):
            try:
                dynamic_images = json.loads(html.unescape(raw_url))
                raw_candidates = dynamic_images.keys()
            except json.JSONDecodeError:
                raw_candidates = []
        else:
            raw_candidates = [raw_url]
        for image_url in raw_candidates:
            normalized = normalize_result_url(urljoin(page_url, image_url))
            if not normalized.startswith(("http://", "https://")) or normalized in seen:
                continue
            seen.add(normalized)
            candidates.append(
                {
                    "url": normalized,
                    "image": normalized,
                    "page_url": page_url,
                    "title": page_title,
                    "source": source_name,
                    "alt": "",
                    "pre_score": score_image_result(
                        product_name,
                        {"url": normalized, "image": normalized, "page_url": page_url, "title": page_title},
                        source_name,
                    )
                    + candidate_url_quality(normalized),
                }
            )
    return sorted(candidates, key=lambda item: item.get("pre_score", 0), reverse=True)[: limits["candidates"]]


def search_images(query, source_name, product_name, limits, required_hosts=()):
    candidates = []
    try:
        with DDGS() as ddgs:
            results = ddgs.images(
                query,
                max_results=limits["image_results"],
                safesearch="moderate",
                size=None,
                type_image="photo",
            )
    except Exception as exc:
        log_image_search(
            "search_failed",
            product=product_name,
            source=source_name,
            query=query,
            error=str(exc),
        )
        return []

    for result in results:
        image_url = normalize_result_url(result.get("image") or result.get("thumbnail"))
        if not image_url.startswith(("http://", "https://")):
            continue
        if required_hosts and not host_matches(image_url, required_hosts):
            continue
        candidate = {
            "url": image_url,
            "image": image_url,
            "thumbnail": result.get("thumbnail"),
            "title": result.get("title") or "",
            "source": source_name,
            "page_url": normalize_result_url(result.get("url") or result.get("source") or ""),
        }
        candidate["pre_score"] = score_image_result(product_name, candidate, source_name) + candidate_url_quality(image_url)
        candidates.append(candidate)

    log_image_search(
        "search_completed",
        product=product_name,
        source=source_name,
        query=query,
        success=bool(candidates),
        candidate_count=len(candidates),
    )
    return candidates


def search_web_pages(query, source_plan, product_name, limits):
    source_name = source_plan["name"]
    domains = source_plan.get("domains", ())
    exclude_domains = source_plan.get("exclude_domains", ())
    max_pages = min(source_plan.get("max_pages", limits["pages"]), limits["pages"])
    page_urls = []

    try:
        with DDGS() as ddgs:
            results = ddgs.text(query, max_results=limits["web_results"], safesearch="moderate")
    except Exception as exc:
        log_image_search(
            "search_failed",
            product=product_name,
            source=source_name,
            query=query,
            error=str(exc),
        )
        return []

    for result in results:
        page_url = normalize_result_url(result.get("href") or result.get("url"))
        if not page_url.startswith(("http://", "https://")):
            continue
        if domains and not host_matches(page_url, domains):
            continue
        if exclude_domains and host_matches(page_url, exclude_domains):
            continue
        if page_url not in page_urls:
            page_urls.append(page_url)
        if len(page_urls) >= max_pages:
            break

    candidates = []
    for page_url in page_urls:
        page_candidates = extract_images_from_page(page_url, product_name, source_name, limits)
        candidates.extend(page_candidates)

    log_image_search(
        "search_completed",
        product=product_name,
        source=source_name,
        query=query,
        success=bool(candidates),
        page_count=len(page_urls),
        candidate_count=len(candidates),
    )
    return candidates


def dedupe_candidates(candidates):
    deduped = {}
    for candidate in candidates:
        url = candidate.get("url")
        if not url:
            continue
        key = canonical_image_key(url)
        score = candidate.get("pre_score", 0)
        existing = deduped.get(key)
        existing_score = (existing or {}).get("pre_score", 0)
        if existing is None or score > existing_score:
            deduped[key] = candidate
    return sorted(deduped.values(), key=lambda item: item.get("pre_score", 0), reverse=True)


def collect_candidates_for_source(source_plan, product_name, query_variants, limits):
    source_candidates = []
    source_name = source_plan["name"]
    for query_variant in query_variants:
        for template in source_plan["queries"]:
            query = source_query_text(template, query_variant, product_name)
            log_image_search(
                "search_started",
                product=product_name,
                source=source_name,
                query=query,
            )
            if source_plan["mode"] == "images":
                candidates = search_images(
                    query,
                    source_name,
                    product_name,
                    limits,
                    required_hosts=source_plan.get("required_hosts", ()),
                )
            else:
                candidates = search_web_pages(query, source_plan, product_name, limits)
            source_candidates.extend(candidates)
            if candidates:
                return dedupe_candidates(source_candidates)[: limits["candidates"]]
    return dedupe_candidates(source_candidates)[: limits["candidates"]]


def select_best_downloaded_candidate(product_name, title_hash, candidates, limits):
    scored = []
    for index, candidate in enumerate(candidates[: limits["downloads"]], start=1):
        suffix = f"_{index}"
        try:
            filename, metadata, byte_size = download_image_to_cache(
                candidate["url"],
                title_hash,
                suffix=suffix,
                timeout=limits["timeout"],
            )
            score = score_downloaded_image(product_name, candidate, metadata, byte_size)
            scored.append(
                {
                    "filename": filename,
                    "source_url": candidate["url"],
                    "source": candidate.get("source", ""),
                    "score": score,
                    "metadata": metadata,
                }
            )
            log_image_search(
                "candidate_downloaded",
                product=product_name,
                source=candidate.get("source", ""),
                image_url=candidate["url"],
                success=True,
                width=metadata["width"],
                height=metadata["height"],
                score=round(score, 3),
            )
        except Exception as exc:
            log_image_search(
                "candidate_download_failed",
                product=product_name,
                source=candidate.get("source", ""),
                image_url=candidate.get("url", ""),
                success=False,
                error=str(exc),
            )

    if not scored:
        return None

    best = max(scored, key=lambda item: item["score"])
    for item in scored:
        if item["filename"] != best["filename"]:
            (IMAGE_CACHE_DIR / item["filename"]).unlink(missing_ok=True)

    final_filename = f"{title_hash}{Path(best['filename']).suffix}"
    final_path = IMAGE_CACHE_DIR / final_filename
    if best["filename"] != final_filename:
        (IMAGE_CACHE_DIR / best["filename"]).replace(final_path)
        best["filename"] = final_filename
    return best


def search_image_candidates(product_name):
    candidates = []
    query_variants = build_search_queries(product_name)
    limits = image_search_limits(deep_search=True)
    for source_plan in limits["source_plans"]:
        source_candidates = collect_candidates_for_source(source_plan, product_name, query_variants, limits)
        candidates.extend(source_candidates)
    ranked = sorted(
        dedupe_candidates(candidates),
        key=lambda item: score_image_result(product_name, item, item.get("source", "")),
        reverse=True,
    )
    return [item["url"] for item in ranked[:40]]


def search_image_url(product_name):
    candidates = search_image_candidates(product_name)
    return candidates[0] if candidates else None


def get_cached_image(title_hash):
    with get_db() as conn:
        row = conn.execute(
            "SELECT image_filename, image_source_url FROM image_cache WHERE title_hash = ?",
            (title_hash,),
        ).fetchone()
    if not row:
        return None
    if (IMAGE_CACHE_DIR / row["image_filename"]).exists():
        return row
    return None


def save_cache_entry(title_hash, product_name, filename, source_url):
    timestamp = now_iso()
    with get_db() as conn:
        conn.execute(
            """
            INSERT INTO image_cache (
                title_hash, product_name, image_filename, image_source_url, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(title_hash) DO UPDATE SET
                product_name = excluded.product_name,
                image_filename = excluded.image_filename,
                image_source_url = excluded.image_source_url,
                updated_at = excluded.updated_at
            """,
            (title_hash, product_name, filename, source_url, timestamp, timestamp),
        )


def get_product(product_id):
    with get_db() as conn:
        return conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()


def save_manual_image_upload(product, image_file):
    ext = Path(image_file.filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        raise ValueError("Use a JPG, PNG, WEBP, or GIF image.")
    filename = f"manual_{product['title_hash']}_{int(time.time())}{ext}"
    image_file.save(IMAGE_CACHE_DIR / filename)
    save_cache_entry(product["title_hash"], product["product_name"], filename, "manual upload")
    return filename, "manual upload"


def save_manual_image_url(product, image_url):
    if not image_url.startswith(("http://", "https://")):
        raise ValueError("Image URL must start with http:// or https://")
    filename = download_image(image_url, f"manual_{product['title_hash']}_{int(time.time())}")
    save_cache_entry(product["title_hash"], product["product_name"], filename, image_url)
    return filename, image_url


def find_or_download_product_image(product_name, force_refresh=False, deep_search=False):
    title_hash = product_hash(product_name)
    if not force_refresh:
        cached = get_cached_image(title_hash)
        if cached:
            return cached["image_filename"], cached["image_source_url"]

    deep_search = deep_search or force_refresh
    limits = image_search_limits(deep_search=deep_search)
    log_image_search(
        "product_started",
        product=product_name,
        title_hash=title_hash,
        force_refresh=force_refresh,
        deep_search=deep_search,
    )
    query_variants = build_search_queries(product_name)
    for source_plan in limits["source_plans"]:
        source_name = source_plan["name"]
        candidates = collect_candidates_for_source(source_plan, product_name, query_variants, limits)
        if not candidates:
            log_image_search("source_exhausted", product=product_name, source=source_name, success=False)
            continue

        candidates = sorted(
            candidates,
            key=lambda item: score_image_result(product_name, item, source_name),
            reverse=True,
        )
        best = select_best_downloaded_candidate(product_name, title_hash, candidates, limits)
        if best:
            save_cache_entry(title_hash, product_name, best["filename"], best["source_url"])
            log_image_search(
                "product_completed",
                product=product_name,
                success=True,
                selected_image_url=best["source_url"],
                final_source=best["source"],
                width=best["metadata"]["width"],
                height=best["metadata"]["height"],
            )
            return best["filename"], best["source_url"]
        log_image_search("source_exhausted", product=product_name, source=source_name, success=False)

    log_image_search(
        "product_completed",
        product=product_name,
        success=False,
        selected_image_url=None,
        final_source=None,
    )
    return PLACEHOLDER_IMAGE, None


def create_import_job(job_id, original_filename, stored_filename):
    timestamp = now_iso()
    with get_db() as conn:
        conn.execute(
            """
            INSERT INTO import_jobs (
                id, original_filename, stored_filename, status, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (job_id, original_filename, stored_filename, "queued", timestamp, timestamp),
        )


def update_import_job(job_id, **fields):
    if not fields:
        return
    fields["updated_at"] = now_iso()
    assignments = ", ".join(f"{key} = ?" for key in fields)
    values = list(fields.values())
    values.append(job_id)
    with get_db() as conn:
        conn.execute(f"UPDATE import_jobs SET {assignments} WHERE id = ?", values)


def get_import_job(job_id):
    with get_db() as conn:
        return conn.execute("SELECT * FROM import_jobs WHERE id = ?", (job_id,)).fetchone()


def import_excel_file(path, job_id=None):
    df = pd.read_excel(path)
    product_col = find_column(df.columns, "Product Name")
    price_columns = detect_price_columns(df.columns)
    if product_col is None:
        raise ValueError("The Excel file must contain a 'Product Name' column.")

    imported = 0
    skipped = 0
    processed = 0
    timestamp = now_iso()
    total = len(df.index)

    if job_id:
        update_import_job(job_id, status="running", total_count=total)

    for _, row in df.iterrows():
        product_name = normalize_product_name(row.get(product_col))
        if not product_name:
            skipped += 1
            processed += 1
            if job_id:
                update_import_job(
                    job_id,
                    processed_count=processed,
                    skipped_count=skipped,
                    current_product="Skipping blank product name",
                )
            continue

        price = get_unit_price_text(row, price_columns)
        title_hash = product_hash(product_name)
        if job_id:
            update_import_job(
                job_id,
                processed_count=processed,
                imported_count=imported,
                skipped_count=skipped,
                current_product=product_name[:240],
            )
        image_filename, source_url = find_or_download_product_image(product_name)
        with get_db() as conn:
            conn.execute(
                """
                INSERT INTO products (
                    product_name, price, image_filename, image_source_url, title_hash, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (product_name, price, image_filename, source_url, title_hash, timestamp, timestamp),
            )
        imported += 1
        processed += 1
        if job_id:
            update_import_job(
                job_id,
                processed_count=processed,
                imported_count=imported,
                skipped_count=skipped,
                current_product=product_name[:240],
            )
    return imported, skipped


def run_import_job(job_id, stored_path):
    job = get_import_job(job_id)
    if not job:
        return

    try:
        imported, skipped = import_excel_file(stored_path, job_id=job_id)
        with get_db() as conn:
            conn.execute(
                """
                INSERT INTO upload_batches (
                    original_filename, stored_filename, imported_count, skipped_count, created_at
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (job["original_filename"], job["stored_filename"], imported, skipped, now_iso()),
            )
        update_import_job(
            job_id,
            status="complete",
            imported_count=imported,
            skipped_count=skipped,
            current_product="Import complete",
            completed_at=now_iso(),
        )
    except Exception as exc:
        update_import_job(
            job_id,
            status="error",
            error_message=str(exc),
            current_product="Import stopped",
            completed_at=now_iso(),
        )


def start_import_thread(job_id, stored_path):
    worker = threading.Thread(target=run_import_job, args=(job_id, stored_path), daemon=True)
    worker.start()


def open_browser_after_start():
    time.sleep(1)
    webbrowser.open("http://127.0.0.1:5000")


def product_query(search, page, per_page):
    offset = (page - 1) * per_page
    where = ""
    params = []
    if search:
        where = "WHERE product_name LIKE ?"
        params.append(f"%{search}%")

    with get_db() as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM products {where}", params).fetchone()[0]
        products = conn.execute(
            f"""
            SELECT * FROM products
            {where}
            ORDER BY id DESC
            LIMIT ? OFFSET ?
            """,
            [*params, per_page, offset],
        ).fetchall()
    return products, total


def all_products(search=""):
    where = ""
    params = []
    if search:
        where = "WHERE product_name LIKE ?"
        params.append(f"%{search}%")
    with get_db() as conn:
        return conn.execute(
            f"""
            SELECT product_name, price, image_filename, image_source_url, created_at, updated_at
            FROM products
            {where}
            ORDER BY id DESC
            """,
            params,
        ).fetchall()


def build_catalog_workbook(products):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    headers = ["Product Name", "Price", "Image File", "Image Source", "Created At", "Updated At"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="116B5F")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for product in products:
        sheet.append(
            [
                product["product_name"],
                product["price"],
                product["image_filename"],
                product["image_source_url"],
                product["created_at"],
                product["updated_at"],
            ]
        )

    widths = [55, 16, 28, 60, 22, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@app.context_processor
def inject_helpers():
    def image_url(filename):
        if not filename or filename == PLACEHOLDER_IMAGE:
            return url_for("static", filename=f"img/{PLACEHOLDER_IMAGE}")
        return url_for("cached_image", filename=filename)

    return {"image_url": image_url}


@app.route("/")
def index():
    search = request.args.get("q", "").strip()
    page = max(request.args.get("page", 1, type=int), 1)
    per_page = min(max(request.args.get("per_page", 24, type=int), 12), 96)
    products, total = product_query(search, page, per_page)
    total_pages = max((total + per_page - 1) // per_page, 1)
    return render_template(
        "index.html",
        products=products,
        total=total,
        search=search,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
    )


@app.route("/export/catalog.xlsx")
def export_catalog():
    search = request.args.get("q", "").strip()
    products = all_products(search)
    output = build_catalog_workbook(products)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        output,
        as_attachment=True,
        download_name=f"inventory_catalog_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/upload", methods=["GET", "POST"])
def upload():
    if request.method == "GET":
        return render_template("upload.html")

    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("Choose an Excel .xlsx file first.", "error")
        return redirect(url_for("upload"))
    if not allowed_excel_file(file.filename):
        flash("Only .xlsx files are supported.", "error")
        return redirect(url_for("upload"))

    safe_name = secure_filename(file.filename)
    stored_name = f"{int(time.time())}_{safe_name}"
    stored_path = UPLOAD_DIR / stored_name
    file.save(stored_path)

    job_id = uuid.uuid4().hex
    create_import_job(job_id, safe_name, stored_name)
    start_import_thread(job_id, stored_path)
    return redirect(url_for("import_progress", job_id=job_id))


@app.route("/import/<job_id>")
def import_progress(job_id):
    job = get_import_job(job_id)
    if not job:
        abort(404)
    return render_template("import_progress.html", job=job)


@app.route("/api/imports/<job_id>")
def import_status(job_id):
    job = get_import_job(job_id)
    if not job:
        return jsonify({"error": "Import job not found"}), 404
    total = job["total_count"] or 0
    processed = job["processed_count"] or 0
    percent = round((processed / total) * 100) if total else 0
    return jsonify(
        {
            "id": job["id"],
            "status": job["status"],
            "total_count": total,
            "processed_count": processed,
            "imported_count": job["imported_count"] or 0,
            "skipped_count": job["skipped_count"] or 0,
            "current_product": job["current_product"] or "",
            "error_message": job["error_message"] or "",
            "percent": percent,
            "catalog_url": url_for("index"),
        }
    )


@app.route("/admin")
def admin():
    search = request.args.get("q", "").strip()
    page = max(request.args.get("page", 1, type=int), 1)
    products, total = product_query(search, page, 20)
    total_pages = max((total + 19) // 20, 1)
    with get_db() as conn:
        cache_count = conn.execute("SELECT COUNT(*) FROM image_cache").fetchone()[0]
    return render_template(
        "admin.html",
        products=products,
        total=total,
        search=search,
        page=page,
        total_pages=total_pages,
        cache_count=cache_count,
    )


@app.route("/admin/product/<int:product_id>/replace-image", methods=["POST"])
def replace_image(product_id):
    image = request.files.get("image_file")
    if not image or not image.filename:
        flash("Choose an image file first.", "error")
        return redirect(url_for("admin"))

    product = get_product(product_id)
    if not product:
        abort(404)

    try:
        filename, source = save_manual_image_upload(product, image)
        with get_db() as conn:
            conn.execute(
                """
                UPDATE products
                SET image_filename = ?, image_source_url = ?, updated_at = ?
                WHERE id = ?
                """,
                (filename, source, now_iso(), product_id),
            )
        flash("Image replaced.", "success")
    except ValueError as exc:
        flash(str(exc), "error")

    return redirect(url_for("admin", q=request.form.get("q", ""), page=request.form.get("page", 1)))


@app.route("/admin/product/<int:product_id>/edit", methods=["GET", "POST"])
def edit_product(product_id):
    product = get_product(product_id)
    if not product:
        abort(404)

    if request.method == "GET":
        return render_template("edit_product.html", product=product)

    product_name = normalize_product_name(request.form.get("product_name"))
    price = get_price_text(request.form.get("price"))
    image_url_value = (request.form.get("image_url") or "").strip()
    image_file = request.files.get("image_file")

    if not product_name:
        flash("Product Name cannot be blank.", "error")
        return redirect(url_for("edit_product", product_id=product_id))

    title_hash = product_hash(product_name)
    image_filename = product["image_filename"]
    image_source_url = product["image_source_url"]

    try:
        product_for_image = dict(product)
        product_for_image["product_name"] = product_name
        product_for_image["title_hash"] = title_hash
        if image_file and image_file.filename:
            image_filename, image_source_url = save_manual_image_upload(product_for_image, image_file)
        elif image_url_value:
            image_filename, image_source_url = save_manual_image_url(product_for_image, image_url_value)

        with get_db() as conn:
            conn.execute(
                """
                UPDATE products
                SET product_name = ?, price = ?, title_hash = ?, image_filename = ?,
                    image_source_url = ?, updated_at = ?
                WHERE id = ?
                """,
                (product_name, price, title_hash, image_filename, image_source_url, now_iso(), product_id),
            )
        flash("Product updated.", "success")
        return redirect(url_for("admin"))
    except Exception as exc:
        flash(str(exc), "error")
        return redirect(url_for("edit_product", product_id=product_id))


@app.route("/admin/product/<int:product_id>/refresh-image", methods=["POST"])
def refresh_image(product_id):
    with get_db() as conn:
        product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        if not product:
            abort(404)

    image_filename, source_url = find_or_download_product_image(product["product_name"], force_refresh=True)
    with get_db() as conn:
        conn.execute(
            """
            UPDATE products
            SET image_filename = ?, image_source_url = ?, updated_at = ?
            WHERE id = ?
            """,
            (image_filename, source_url, now_iso(), product_id),
        )
    flash("Image search refreshed using the full product title.", "success")
    return redirect(url_for("admin", q=request.form.get("q", ""), page=request.form.get("page", 1)))


@app.route("/admin/cache/delete", methods=["POST"])
def delete_cache():
    with get_db() as conn:
        rows = conn.execute("SELECT image_filename FROM image_cache").fetchall()
        for row in rows:
            path = IMAGE_CACHE_DIR / row["image_filename"]
            if path.name != PLACEHOLDER_IMAGE:
                path.unlink(missing_ok=True)
        conn.execute("DELETE FROM image_cache")
        conn.execute(
            """
            UPDATE products
            SET image_filename = ?, image_source_url = ?, updated_at = ?
            """,
            (PLACEHOLDER_IMAGE, None, now_iso()),
        )
    flash("Cached images deleted. Products now show the placeholder until refreshed.", "success")
    return redirect(url_for("admin"))


@app.route("/cached-images/<path:filename>")
def cached_image(filename):
    return send_from_directory(IMAGE_CACHE_DIR, filename)


if __name__ == "__main__":
    init_db()
    debug_enabled = os.environ.get("FLASK_DEBUG", "0") == "1"
    if os.environ.get("OPEN_BROWSER", "0") == "1" or getattr(sys, "frozen", False):
        threading.Thread(target=open_browser_after_start, daemon=True).start()
    app.run(debug=debug_enabled, use_reloader=False, host="127.0.0.1", port=5000)
